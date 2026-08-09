"""
modules/parsing.py
Excel / CSV ingestion: classify sheet type, parse rows into list-of-dicts,
skip aggregate/totals rows.

CHANGED (sub-row match detail):
  _classify_subrow_cell() now returns (field_name, match_detail) so the
  transformation journey dialog can show exactly which regex or heuristic
  caused a sub-row cell to be assigned to "Address", "Cause of Loss", etc.
  _enrich_field() stores match_detail and subrow_inferred=True in the claim
  dict so dialogs.py can surface this in Step 2 of the field timeline.
"""

import csv
import os
import re

import openpyxl
import datetime as _dt

from modules.cell_format import format_cell_value_with_fmt
MAX_EXCEL_ROWS = 50_000  # sanity ceiling — a legitimate loss run is nowhere near this

# ── Sheet classifier ──────────────────────────────────────────────────────────

def _loss_run_score(text: str) -> int:
    score = 0
    if any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no", "file num", "file ref",
    ]):
        score += 2
    if any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date", "dol",
    ]):
        score += 1
    if any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ]):
        score += 1
    return score


def _commercial_loss_run_score(text: str) -> int:
    if "policy" in text and ("claim" in text or "incurred" in text):
        return 3
    return 0


def _score_to_confidence(top_score: int, margin_ratio: float) -> float:
    # Confidence depends on BOTH how strong the winning signal is
    # (top_score) and how far ahead it is of the runner-up (margin_ratio).
    # A lone weak signal with nothing to compete against should NOT get
    # near-maximum confidence just because "nothing else scored" -- that
    # was the flaw in scoring purely on margin.
    if top_score >= 3:
        base = 0.65
    elif top_score == 2:
        base = 0.50
    else:
        base = 0.35
    return round(min(0.85, base + 0.20 * margin_ratio), 2)


def classify_sheet(rows) -> tuple[str, float]:
    """
    Returns (sheet_type, confidence) instead of a bare string. Confidence
    is on the same 0-1 scale used by classification_router's keyword
    pre-pass for PDF/TXT/HTML, so every parser in the app now reports
    classification confidence the same way.

    ORDERING NOTE: Check Register's hard trigger runs BEFORE the scored
    Loss Run / Commercial Loss Run competition, exactly as before -- this
    is deliberate. An insurance check register legitimately contains
    Claim Number and Policy Number columns, so if it competed on score
    against those two, a check register with several claim-reference
    columns could out-score its own check-register signal and get
    misclassified as Commercial Loss Run again -- the exact bug this
    ordering was added to prevent. Confidence for Check Register is
    computed from how many of its OWN signals fired, not by racing it
    against the other types.
    """
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)

    if "line of business" in text:
        summary_co_signals = [
            "# claims", "num claims", "number of claims", "claim count",
            "loss ratio", "loss rate", "frequency", "severity",
        ]
        if any(sig in text for sig in summary_co_signals):
            return "SUMMARY", 0.9
        for row in rows[:20]:
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if non_empty and str(non_empty[0]).lower().strip() == "line of business" and len(non_empty) == 1:
                return "SUMMARY", 0.9

    # ── Check Register detection (hard trigger, ordering unchanged) ────────
    has_check_no = any(x in text for x in [
        "check number", "check no.", "check no", "check #", "check num",
        "cheque number", "cheque no",
    ])
    has_payee = "payee" in text
    has_register_ledger = any(x in text for x in [
        "running total", "running balance", "payment/debit", "deposit/credit", "balance",
    ])
    if has_check_no and (has_payee or has_register_ledger):
        signal_count = sum([has_check_no, has_payee, has_register_ledger])
        confidence   = min(0.85, 0.55 + 0.10 * signal_count)
        return "CHECK_REGISTER", round(confidence, 2)

    # ── Loss Run vs. Commercial Loss Run: scored competition ────────────────
    scores = {
        "LOSS_RUN":            _loss_run_score(text),
        "COMMERCIAL_LOSS_RUN": _commercial_loss_run_score(text),
    }
    ranked = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    top_type, top_score = ranked[0]
    second_score = ranked[1][1]

    if top_score == 0:
        return "UNKNOWN", 0.0

    margin_ratio = (top_score - second_score) / max(top_score, 1)
    return top_type, _score_to_confidence(top_score, margin_ratio)


# ── Legacy-layout detector ────────────────────────────────────────────────────

def _is_legacy_print_layout(rows: list) -> bool:
    for row in rows:
        non_empty = [c for c in row if c is not None]
        if non_empty and all(str(c).strip() == "----------" for c in non_empty):
            return True

    for i in range(min(20, len(rows) - 1)):
        r1_vals = [str(c).strip() for c in rows[i] if c is not None]
        r2_vals = [str(c).strip() for c in rows[i + 1] if c is not None]
        if len(r2_vals) >= 5 and len(r1_vals) >= 2:
            r1_filled = sum(1 for c in rows[i] if c)
            r2_filled = sum(1 for c in rows[i + 1] if c)
            if r2_filled > r1_filled * 1.5 and r1_filled >= 2:
                combined = " ".join(r1_vals + r2_vals).lower()
                if ("file" in combined or "claim" in combined) and (
                    "paid" in combined or "incurred" in combined or "outstanding" in combined
                ):
                    return True
    return False


def _find_legacy_header_rows(rows: list) -> tuple[int, int] | None:
    for i in range(min(25, len(rows) - 1)):
        r1 = rows[i]
        r2 = rows[i + 1]
        r1_filled = sum(1 for c in r1 if c)
        r2_filled = sum(1 for c in r2 if c)
        if r2_filled < 4:
            continue
        combined = " ".join(
            str(c).lower() for c in list(r1) + list(r2) if c
        )
        if ("file" in combined or "claim" in combined or "assured" in combined) and (
            "paid" in combined or "outstanding" in combined or "incurred" in combined
        ):
            if r1_filled >= 2:
                return (i, i + 1)
            if r2_filled >= 5:
                return (i + 1, i + 1)
    return None


def _merge_two_header_rows(row1: list, row2: list) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for g, s in zip(row1, row2):
        g_s = str(g).strip() if g else ""
        s_s = str(s).strip() if s else ""
        if g_s and s_s and g_s.upper() != s_s.upper():
            name = f"{g_s} {s_s}"
        elif s_s:
            name = s_s
        elif g_s:
            name = g_s
        else:
            name = ""
        if name:
            seen[name] = seen.get(name, 0) + 1
            if seen[name] > 1:
                name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


# ── Sub-row / separator / subtotal detectors ─────────────────────────────────

def _is_separator_row(row_values: list) -> bool:
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if not non_empty:
        return False
    return all(str(c).strip() == "----------" for c in non_empty)


def _is_subtotal_row(row_values: list) -> bool:
    for c in row_values:
        if c is not None and str(c).strip():
            return bool(re.match(r"^total\b", str(c).strip(), re.IGNORECASE))
    return False


def _is_legacy_sub_row(row_values: list, num_cols: int) -> bool:
    if not row_values or row_values[0] is not None:
        return False
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if len(non_empty) == 0 or len(non_empty) > 3:
        return False
    has_addr_or_cause = (
        (len(row_values) > 1 and row_values[1] is not None) or
        (len(row_values) > 3 and row_values[3] is not None)
    )
    return has_addr_or_cause


# ── Smart sub-row cell classifier ─────────────────────────────────────────────

_ADDRESS_PAT = re.compile(
    r"""
    ^(
        \d+\s+\w.*            # "391 MAIN ST"
      | P\.?O\.?\s*BOX\s+\d   # "PO BOX 443"
      | \d+[-/]\d+\s+\w.*     # "12-14 ELM RD"
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

_ADDRESS_SUFFIX = re.compile(
    r"\b(st|street|ave|avenue|blvd|boulevard|rd|road|dr|drive|"
    r"ln|lane|ct|court|pl|place|way|cir|circle|hwy|highway|"
    r"pkwy|parkway|terr|ter|loop|trail|trl|run|box|suite|ste|"
    r"apt|unit|floor|fl)\b",
    re.IGNORECASE,
)

_CITY_STATE_ZIP_PAT = re.compile(
    r"""
    ^(
        [A-Za-z\s]{2,30},?\s+[A-Z]{2}\s+\d{5}(-\d{4})?   # City, ST 12345
      | [A-Za-z\s]{2,30},?\s+[A-Z]{2}$                     # City, ST
      | \d{5}(-\d{4})?$                                     # ZIP only
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

_CAUSE_OF_LOSS_PAT = re.compile(
    r"\b(fire|flood|wind|windstorm|hail|storm|tornado|hurricane|"
    r"tropical\s+storm|water\s+damage|water\s+intrusion|theft|"
    r"vandalism|slip|fall|trip|collision|accident|explosion|"
    r"lightning|freeze|ice|snow|earthquake|sinkhole|mold|"
    r"liability|negligence|assault|discrimination|wrongful|"
    r"product\s+liability|premises|auto|vehicle|medical|workers|"
    r"comp|injury|glass|burst\s+pipe|pipe\s+burst|roof|damage)\b",
    re.IGNORECASE,
)

_NAME_PAT = re.compile(
    r"^[A-Z][A-Za-z'-]+(\s+[A-Z][A-Za-z'-]+){1,4}$"
)


def _col_letter(col_index: int) -> str:
    result = ""
    n = col_index + 1
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


# ── CHANGED: returns (field_name, match_detail) ───────────────────────────────
def _classify_subrow_cell(value: str) -> tuple[str, str]:
    """
    Classify a single sub-row cell value into a semantic field name.

    Returns (field_name, match_detail) where match_detail is a human-readable
    description of which regex / heuristic caused the classification.
    This detail is forwarded all the way to the transformation journey dialog
    so reviewers can see exactly why a field was inferred (e.g. "regex: peril
    keyword matched — 'Tropical Storm'").

    field_name is one of:
      "Address"        – looks like a street address line
      "City State Zip" – looks like city/state/zip continuation
      "Cause of Loss"  – matches known peril / event vocabulary
      "Claimant Name"  – looks like a person/company name
      "Unknown"        – cannot be confidently classified
    """
    v = str(value).strip()
    if not v:
        return "Unknown", "empty value"

    if _ADDRESS_PAT.match(v):
        return "Address", "regex: street-number + name pattern (e.g. '391 MAIN ST', 'PO BOX …')"
    if _ADDRESS_SUFFIX.search(v) and re.search(r'\d', v):
        suffix_m = _ADDRESS_SUFFIX.search(v)
        suffix   = suffix_m.group(0) if suffix_m else ""
        return "Address", f"regex: street-suffix keyword '{suffix.upper()}' + digit present"
    if _CITY_STATE_ZIP_PAT.match(v):
        return "City State Zip", "regex: City ST 00000 / two-letter state / ZIP-only pattern"
    m = _CAUSE_OF_LOSS_PAT.search(v)
    if m:
        peril = m.group(0)
        return "Cause of Loss", f"regex: peril keyword matched — \"{peril}\""
    if _NAME_PAT.match(v) and not re.search(r'\d', v):
        return "Claimant Name", "regex: proper-noun name pattern (Title-cased words, no digits)"
    return "Unknown", "no pattern matched — stored verbatim"


# ── CHANGED: result tuple is now (value, col, match_detail) ──────────────────
def _infer_subrow_fields(raw_row: list) -> dict[str, tuple[str, int, str]]:
    """
    Given a raw sub-row (list of cell values), return a dict mapping
    inferred field names → (value, 1-based col index, match_detail).

    match_detail is the human-readable description of which regex/heuristic
    produced the classification — forwarded to _enrich_field() so the
    transformation journey dialog can show it.
    """
    result: dict[str, tuple[str, int, str]] = {}
    type_count: dict[str, int] = {}

    for c_idx, val in enumerate(raw_row):
        if val is None or not str(val).strip():
            continue

        val_s                  = str(val).strip()
        field_type, match_detail = _classify_subrow_cell(val_s)

        if field_type == "Unknown":
            field_type   = f"SubRow_{_col_letter(c_idx)}"
            match_detail = f"unclassified value in col {_col_letter(c_idx)} — stored verbatim"

        type_count[field_type] = type_count.get(field_type, 0) + 1
        if type_count[field_type] > 1:
            field_type = f"{field_type}_{type_count[field_type]}"

        result[field_type] = (val_s, c_idx + 1, match_detail)   # 1-based col

    return result


def _enrich_from_subrow(
    claim: dict,
    raw_row: list,
    r_idx: int,
) -> None:
    """
    Smart replacement for the hardcoded col-B/col-D sub-row enrichment.

    Infers field names from sub-row cell values using pattern matching and
    calls _enrich_field() for each one.  match_detail is forwarded so
    the transformation journey can show how each sub-row field was inferred.
    """
    inferred = _infer_subrow_fields(raw_row)
    for field_name, (value, excel_col, match_detail) in inferred.items():
        _enrich_field(
            claim, field_name, value,
            excel_row=r_idx, excel_col=excel_col,
            match_detail=match_detail,
        )


# ── Aggregate-row detection ───────────────────────────────────────────────────

_AGGREGATE_PATTERNS = re.compile(
    r"^(total|totals|grand\s*total|subtotal|aggregate|summary|sum|report\s*(date|total|summary)|"
    r"all\s+adjusters|ytd\s+total|period\s+total|fiscal\s+total|portfolio\s+total|"
    r"TOTALS_AGGREGATE|SUMMARY_FLIBBER|AGGREGATE_ZORP|SUMMARY_ZORP)",
    re.IGNORECASE,
)
_AGGREGATE_EXTRA = re.compile(
    r"(aggregate|zorp|flibber|summary|zoop|gorp|totals?_|_total|report_date|all_adjuster)",
    re.IGNORECASE,
)


def _is_aggregate_row(row_values: list) -> bool:
    non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return False
    first_val = non_empty[0]
    if _AGGREGATE_PATTERNS.match(first_val):
        return True
    if _AGGREGATE_EXTRA.search(first_val):
        return True
    first_tokens     = re.split(r"[_\s]+", first_val.lower())
    aggregate_tokens = {"total", "totals", "aggregate", "summary", "subtotal", "grand", "portfolio", "report"}
    if len(first_tokens) >= 2 and any(t in aggregate_tokens for t in first_tokens):
        return True
    for v in non_empty[:6]:
        if re.match(
            r"(total\s+claims|report\s+date|all\s+adjusters|open:\s*\d|pend:\s*\d|open:\d)",
            str(v), re.IGNORECASE,
        ):
            return True
    nums = [float(v) for v in row_values if isinstance(v, (int, float))]
    if nums and len(nums) >= 3 and all(n > 50_000 for n in nums):
        is_claim_id = (
            re.match(r"^[A-Z]{2,5}[-_][A-Z]{0,3}\d{3,}", first_val, re.IGNORECASE)
            or re.match(r"^\d{4,}$", first_val.strip())
        )
        if not is_claim_id:
            return True
    return False


# ── Sheet title / metadata extractor ─────────────────────────────────────────

_LABEL_ALIASES: dict[str, str] = {
    "prepared for":     "Reinsurer",
    "reinsurer":        "Reinsurer",
    "prepared by":      "TPA Name",
    "treaty":           "Treaty",
    "program":          "Treaty",
    "policy":           "Policy Number",
    "cedant":           "Cedant",
    "ceding company":   "Cedant",
    "insurer":          "Cedant",
    "valuation date":   "Valuation Date",
    "valuation":        "Valuation Date",
    "as of":            "Valuation Date",
    "report date":      "Report Date",
    "report generated": "Report Date",
    "effective date":   "Effective Date",
    "policy number":    "Policy Number",
    "policy no":        "Policy Number",
    "policy #":         "Policy Number",
    "insured":          "Insured Name",
    "named insured":    "Insured Name",
    "line of business": "Line of Business",
    "lob":              "Line of Business",
    "coverage":         "Coverage Type",

     # ── NEW: labels this file's banner actually uses ──────────────────────────
    "period ended": "Valuation Date", "period ending": "Valuation Date",
    "period end": "Valuation Date", "as at": "Valuation Date",
    "underwriting period": "Underwriting Period",
    "policy period": "Policy Period", "coverage period": "Policy Period",
    "contract number": "Contract Number", "contract no": "Contract Number",
    "contract #": "Contract Number", "contract": "Contract Number",
    "carrier": "Carrier", "carrier name": "Carrier", "underwriter": "Carrier",
    "assured": "Insured Name", "account": "Insured Name",
    "program year": "Program Year", "claims administrator": "TPA Name",

}

# Labels whose VALUE is a date range get split into two fields, matching the
# naming extract_title_fields() already emits for merged-cell banners.
_RANGE_FIELDS = {
    "Underwriting Period": ("Underwriting Period Start", "Underwriting Period End"),
    "Policy Period":       ("Policy Period Start", "Policy Period End"),
}

_DATE_TOKEN = (
    r"(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4}"
    r"|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}"
    r"|\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}"
    r"|\d{4}-\d{2}-\d{2})"
)

_RANGE_PAT = re.compile(
    rf"^{_DATE_TOKEN}\s*(?:to|through|thru|[\u2013\u2014\-])\s*{_DATE_TOKEN}$", re.IGNORECASE
)

_COMPANY_SUFFIX = re.compile(
    r"\b(llc|l\.l\.c|inc|inc\.|incorporated|corp|corporation|co\.|co|company|ltd|limited|"
    r"lp|llp|plc|holdings|group|partners|properties|associates|enterprises|"
    r"ins\.?|insurance|assurance|underwriters|mgmt|management|trust|estates)\b",
    re.IGNORECASE,
)

# Same vocabulary as extract_title_fields()'s lob_map, so both paths agree.
_LOB_MAP = [
    (r"workers[\u2019'\s\-]*comp(?:ensation)?", "Workers Compensation"),
    (r"commercial\s+general\s+liability",       "Commercial General Liability"),
    (r"commercial\s+auto(?:mobile)?",           "Commercial Auto"),
    (r"commercial\s+prop(?:erty)?",             "Commercial Property"),
    (r"professional\s+liability",               "Professional Liability"),
    (r"general\s+liability",                    "General Liability"),
    (r"\bumbrella\b",                           "Umbrella"),
    (r"\bexcess\s+casualty\b",                  "Excess Casualty"),
    (r"\binland\s+marine\b",                    "Inland Marine"),
    (r"\bcyber\b",                              "Cyber"),
]

_NOTE_PAT   = re.compile(r"^\(.*\)$|^(see|note|source|continued)\b", re.IGNORECASE)
_LABELISH   = re.compile(r"^[A-Za-z][A-Za-z0-9 &/#.'\-]{1,38}$")
_VALUE_LIKE = re.compile(r"^[\dA-Z]")


def _canonical_label(raw: str) -> str | None:
    key = re.sub(r"\s+", " ", str(raw).strip().rstrip(":").lower())
    return _LABEL_ALIASES.get(key)


def _try_inline_kv(cell_text: str) -> list[tuple[str, str]]:
    pairs = []
    for seg in re.split(r'\s{3,}|\|', str(cell_text)):
        m = re.match(r'^([A-Za-z][^:]{0,40}):\s*(.+)$', seg.strip())
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
    return pairs


def _looks_like_label(text: str) -> bool:
    """A banner label cell. Accepts a known alias, a trailing colon, OR a short
    ALL-CAPS caption with no digits -- which is how print-style loss runs write
    'PERIOD ENDED' in the cell to the left of its value."""
    t = str(text).strip()
    if not t or len(t.split()) > 5 or not _LABELISH.match(t):
        return False
    if _canonical_label(t) or t.endswith(":"):
        return True
    letters = [ch for ch in t if ch.isalpha()]
    return bool(letters) and all(ch.isupper() for ch in letters) and not re.search(r"\d", t)


def _classify_banner(text: str) -> tuple[str, str] | None:
    """Unlabelled banner line -> (field, value). '_COMPANY_' means 'an org name,
    caller decides whether it's the carrier or the insured by position'."""
    t = str(text).strip()
    if not t or _NOTE_PAT.match(t):
        return None
    for pat, lob in _LOB_MAP:
        if re.search(pat, t, re.IGNORECASE):
            return "Line of Business", lob
    if _COMPANY_SUFFIX.search(t):
        return "_COMPANY_", t
    return None


def extract_sheet_title_kvs(
    raw_rows: list,
    cell_rows: list,
    header_row_idx: int | None,
    sheet_name: str,
) -> dict:
    """
    Three passes per banner row, instead of the old single-cell special case:
      1. "LABEL: value" inside one cell            (CONTRACT NUMBER:PLIC/25)
      2. label cell + value cell to its right      (PERIOD ENDED | 12/31/2025)
      3. unlabelled caption lines                  (AR LLC, Commercial Property)
    A cell consumed by an earlier pass is never re-read by a later one, so the
    same text can't land in two fields.
    """
    scan_limit = header_row_idx if header_row_idx is not None else min(15, len(raw_rows))
    found: dict = {}
    company_seen = 0

    def _disp(r_idx: int, c_idx: int, fallback) -> str:
        """Displayed value -- honours the cell's number format so a real date
        reads '12/31/2025', not '2025-12-31 00:00:00'. This is what cell_rows
        was passed in for; the old body never touched it."""
        try:
            return format_cell_value_with_fmt(cell_rows[r_idx][c_idx])
        except Exception:
            if isinstance(fallback, (_dt.datetime, _dt.date)):
                return fallback.strftime("%m/%d/%Y")
            return str(fallback).strip()

    def _store(canonical: str, value: str, excel_row: int, excel_col: int,
               source: str = "title_kv") -> None:
        value = str(value).strip()
        if not canonical or not value:
            return
        if canonical in _RANGE_FIELDS:
            m = _RANGE_PAT.match(value)
            if m:
                s_field, e_field = _RANGE_FIELDS[canonical]
                _store(s_field, m.group(1), excel_row, excel_col, source)
                _store(e_field, m.group(2), excel_row, excel_col, source)
                return
        if canonical in found:
            return
        found[canonical] = {
            "value":     value,
            "original":  value,
            "modified":  value,
            "source":    source,
            "excel_row": excel_row,
            "excel_col": excel_col,
        }

    for r_idx, row in enumerate(raw_rows[:scan_limit]):
        excel_row = r_idx + 1
        non_empty = [(c, v) for c, v in enumerate(row) if v is not None and str(v).strip()]
        if not non_empty:
            continue

        consumed: set[int] = set()

        # ── Pass 1: "LABEL: value" within a single cell ──────────────────────
        for c_idx, val in non_empty:
            val_s = _disp(r_idx, c_idx, val)
            if ":" in val_s and not re.match(r'^\d', val_s):
                for raw_label, raw_value in _try_inline_kv(val_s):
                    canonical = _canonical_label(raw_label)
                    if canonical:
                        _store(canonical, raw_value, excel_row, c_idx + 1)
                        consumed.add(c_idx)

        # ── Pass 2: label cell + adjacent value cell ─────────────────────────
        i = 0
        while i < len(non_empty) - 1:
            c_label_idx, label_val = non_empty[i]
            c_value_idx, value_val = non_empty[i + 1]
            label_s = _disp(r_idx, c_label_idx, label_val)
            value_s = _disp(r_idx, c_value_idx, value_val)

            if c_label_idx in consumed or (":" in label_s and not label_s.endswith(":")):
                i += 1
                continue
            # <=3 columns apart: a label and its value sit side by side. Without
            # this, two unrelated cells at opposite ends of a wide banner row
            # get paired as label/value.
            if (_looks_like_label(label_s)
                    and (c_value_idx - c_label_idx) <= 3
                    and _VALUE_LIKE.match(value_s)):
                canonical = _canonical_label(label_s)
                if canonical:
                    _store(canonical, value_s, excel_row, c_value_idx + 1)
                    consumed.update({c_label_idx, c_value_idx})
                    i += 2
                    continue
            i += 1

        # ── Pass 3: unlabelled caption lines ─────────────────────────────────
        # A caption is short and wide. A row with 4+ populated cells is a header
        # or data row that leaked inside the scan window -- legacy two-row
        # headers do exactly this -- so never read it as prose.
        if len(non_empty) >= 4:
            continue

        for c_idx, val in non_empty:
            if c_idx in consumed:
                continue
            val_s = _disp(r_idx, c_idx, val)
            if ":" in val_s or re.match(r'^[\d$,()\-\.]+$', val_s):
                continue

            if r_idx == 0 and "TPA Name" not in found:
                tpa = re.split(r'\s*[\u2014\u2013]\s*', val_s)[0].strip()
                if ' - ' in tpa:
                    head, tail = tpa.split(' - ', 1)
                    if re.search(r'\b(report|run|detail|summary|schedule|listing)\b',
                                 tail, re.IGNORECASE):
                        tpa = head.strip()
                _store("TPA Name", tpa, excel_row, c_idx + 1, "title_banner")
                if _COMPANY_SUFFIX.search(tpa):
                    company_seen += 1
                    _store("Carrier", tpa, excel_row, c_idx + 1, "title_banner")
                continue

            hit = _classify_banner(val_s)
            if hit and hit[0] == "_COMPANY_":
                # First org name in the banner is the paper; the next one down
                # is the insured. Matches how these reports are laid out.
                company_seen += 1
                _store("Carrier" if company_seen == 1 else "Insured Name",
                       hit[1], excel_row, c_idx + 1, "title_banner")
            elif hit:
                _store(hit[0], hit[1], excel_row, c_idx + 1, "title_banner")
            else:
                lob_m = re.search(
                    r'(?:loss\s+run\s+report|annual\s+loss\s+run)\s*[\u2014\-\u2013]+\s*(.+)',
                    val_s, re.IGNORECASE,
                )
                _store("Sheet Title", lob_m.group(1).strip() if lob_m else val_s,
                       excel_row, c_idx + 1, "title_banner")

    found.setdefault("Sheet Name", {
        "value": sheet_name, "original": sheet_name, "modified": sheet_name,
        "source": "sheet_tab", "excel_row": 0, "excel_col": 0,
    })
    return found

def _warn_uncalculated_formulas(file_path: str, sheet_name: str) -> int:
    """Detects cells with a formula but no cached value (data_only=True
    would silently return None for these). Read-only scan — doesn't
    affect extraction, just surfaces a count for a UI warning. Only
    meaningful for .xlsx; CSV has no formulas."""
    if os.path.splitext(file_path)[1].lower() == ".csv":
        return 0
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values   = openpyxl.load_workbook(file_path, data_only=True)
        ws_f, ws_v  = wb_formulas[sheet_name], wb_values[sheet_name]
        count = sum(
            1 for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows())
            for cf, cv in zip(row_f, row_v)
            if isinstance(cf.value, str) and cf.value.startswith("=") and cv.value is None
        )
        wb_formulas.close(); wb_values.close()
        return count
    except Exception:
        return 0



# ── Main entry point ──────────────────────────────────────────────────────────

def extract_from_excel(
    file_path: str,
    sheet_name: str,
) -> tuple[list, str, dict, float]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN", {}, 0.0
        if len(rows) > MAX_EXCEL_ROWS:                                          # NEW
            raise ValueError(                                                    # NEW
                f"This CSV has {len(rows):,} rows, exceeding the "               # NEW
                f"{MAX_EXCEL_ROWS:,}-row processing limit. Split the file "      # NEW
                f"into smaller batches and re-upload."                           # NEW
            )                                                                    # NEW
        sheet_type, sheet_confidence = classify_sheet(rows)
        claims, sheet_type = parse_rows(sheet_type, rows)
        return claims, sheet_type, {}, sheet_confidence
    else:
        try:                                                                     # NEW
            wb = openpyxl.load_workbook(file_path, data_only=True)               # CHANGED (was un-wrapped)
        except Exception as e:                                                   # NEW
            raise ValueError(                                                    # NEW
                f"Could not open this Excel file — it may be password-"          # NEW
                f"protected, corrupted, or in an unsupported format "            # NEW
                f"({type(e).__name__}: {e})."                                    # NEW
            ) from e                                                             # NEW
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN", {}, 0.0
        if len(raw_rows) > MAX_EXCEL_ROWS:                                       # NEW
            raise ValueError(                                                    # NEW
                f"Sheet '{sheet_name}' has {len(raw_rows):,} rows, exceeding "   # NEW
                f"the {MAX_EXCEL_ROWS:,}-row processing limit. Split the "       # NEW
                f"file and re-upload."                                           # NEW
            )                                                                    # NEW
        sheet_type, sheet_confidence = classify_sheet(raw_rows)

        # Use the same header detection as the row parser so the title scan
        # and data extraction agree on where the header begins.
        if _is_legacy_print_layout(raw_rows):
            legacy_pair = _find_legacy_header_rows(raw_rows)
            hri = legacy_pair[0] if legacy_pair else _find_header_row(raw_rows)
        else:
            hri = _find_header_row(raw_rows)

        title_kvs = extract_sheet_title_kvs(raw_rows, cell_rows, hri, sheet_name)
        claims, sheet_type = parse_rows_with_cells(sheet_type, raw_rows, cell_rows)
        return claims, sheet_type, title_kvs, sheet_confidence

# ── Row parsers ───────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if (
            "claim" in rt or "employee name" in rt or "driver name" in rt
            or "claimant" in rt or "file" in rt
        ) and (
            "date" in rt or "incurred" in rt or "paid" in rt
            or "injury" in rt or "incident" in rt or "amount" in rt or "reserve" in rt
        ):
            return i
        # ── Check Register headers ──────────────────────────────────────────
        if (
            "check number" in rt or "check no" in rt or "check #" in rt
            or "cheque number" in rt or "cheque no" in rt
        ) and (
            "payee" in rt or "date" in rt or "amount" in rt
            or "balance" in rt or "debit" in rt or "credit" in rt
        ):
            return i
    for i, row in enumerate(rows[:20]):          # widened from rows[:5]
        if sum(1 for c in row if c) >= 3:
            return i
    return None


def parse_rows_with_cells(sheet_type: str, rows: list, cell_rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
            r_idx = hri + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data: dict = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_with_cells(sheet_type, rows, cell_rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
        r_idx = hri + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total", "subtotal"] for c in raw_row if c):
            break
        if _is_aggregate_row(raw_row):
            continue
        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_with_cells(
    sheet_type: str, rows: list, cell_rows: list
) -> tuple[list, str]:
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(rows[top_hri])
            ]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(rows[i]) for i in range(len(rows))) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx_rel, (raw_row, cell_row) in enumerate(
        zip(rows[data_start:], cell_rows[data_start:])
    ):
        r_idx = data_start + 1 + r_idx_rel

        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            continue

        if _is_aggregate_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header = headers[c_idx_0]
            if not header:
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }

        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


# ── CHANGED: accepts match_detail kwarg ──────────────────────────────────────
def _enrich_field(
    claim: dict, field_name: str, value: str,
    excel_row: int, excel_col: int,
    match_detail: str = "",
) -> None:
    """Add or update a field in a claim dict if not already set.

    match_detail is stored in the claim dict so the transformation journey
    dialog can surface exactly how a sub-row field was inferred (e.g.
    'regex: peril keyword matched — "Tropical Storm"').
    subrow_inferred=True marks these fields as coming from a legacy sub-row
    rather than a header column.
    """
    if field_name not in claim or not claim[field_name].get("value"):
        claim[field_name] = {
            "value":           value,
            "modified":        value,
            "excel_row":       excel_row,
            "excel_col":       excel_col,
            "subrow_inferred": True,
            "match_detail":    match_detail or "sub-row pattern inference",
        }


# ── CSV / plain parse_rows (no cell objects) ──────────────────────────────────

def parse_rows(sheet_type: str, rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
            if not any(row):
                continue
            if _is_aggregate_row(list(row)):
                continue
            row_data: dict = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx - 1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_plain(sheet_type, rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        if _is_aggregate_row(list(row)):
            continue
        row_data: dict = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx - 1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_plain(sheet_type: str, rows: list) -> tuple[list, str]:
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[top_hri])]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(r) for r in rows) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            continue

        if _is_aggregate_row(list(raw_row)):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx, value in enumerate(raw_row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            if not header:
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


# ─────────────────────────────────────────────────────────────────────────────
# TXT / TRANSCRIPT PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_txt_file(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a plain-text file (FNOL transcript, notes, etc.) into the
    structure expected by run_pdf_intelligence().
 
    Returns a dict with:
      - "full_text"   : entire decoded text  (primary — read by extract_full_text_from_parsed)
      - "pages"       : single-element list with page_num=1 and raw_text set
                        (kept for compatibility with _build_azure_di_index_from_parsed)
    """
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")
 
    text = text.strip()
 
    return {
        "full_text": text,          # ← THIS is what extract_full_text_from_parsed reads
        "pages": [
            {
                "page_num":  1,
                "raw_text":  text,   # ← fallback if full_text key is not found
                "fields":    [],     # no ADI key-value pairs for TXT
                "page_width":  8.5,
                "page_height": 11.0,
            }
        ],
    }   


# ─────────────────────────────────────────────────────────────────────────────
# TXT / TRANSCRIPT PARSER   (appended to modules/parsing.py)
# Strategy: regex-first → LLM fallback if < MIN_REGEX_FIELDS found
# No hardcoded field names.
# ─────────────────────────────────────────────────────────────────────────────
 
import re as _re_txt
import os as _os_txt
import json as _json_txt
 
_TXT_MIN_REGEX_FIELDS = 3
 
# Generic colon-separated  "Label: value"
_RE_TXT_COLON = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{1,54}?)\s*:\s+(.+)$",
    _re_txt.MULTILINE,
)

# Generic equals-separated  "Label = value"
_RE_TXT_EQUALS = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{1,54}?)\s*=\s+(.+)$",
    _re_txt.MULTILINE,
)

# Table-style  "Label    Value"  (2+ spaces between)
_RE_TXT_TABLE = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{2,40})\s{2,}\t*([^\s].+)$",
    _re_txt.MULTILINE,
)
 
 
def _txt_is_plausible_label(label: str) -> bool:
    label = label.strip()
    if not label or len(label) > 55 or len(label.split()) > 7:
        return False
    if not label[0].isalpha():
        return False
    if _re_txt.search(r"\b(is|are|was|were|has|have|had|will|would|should|can|may)\b", label, _re_txt.I):
        return False
    return True
 
 
def _txt_regex_extract(text: str) -> list:
    seen: dict = {}
 
    def _nk(s: str) -> str:
        return _re_txt.sub(r"[\\s_\\-]+", "_", s.strip().lower())
 
    def _add(label: str, value: str, conf: float, raw: str) -> None:
        label = label.strip()
        value = value.strip()
        if not _txt_is_plausible_label(label) or not value or len(value) > 2000:
            return
        if value == value.upper() and len(value.split()) <= 4 and not _re_txt.search(r"\d", value):
            return
        nk = _nk(label)
        if nk not in seen:
            seen[nk] = {
                "field_name": label, "value": value,
                "confidence": conf, "source_text": raw.strip()[:200],
                "bounding_polygon": None, "source_page": 1,
                "page_width": None, "page_height": None,
                "excel_row": 1, "excel_col": None,
            }
 
    for m in _RE_TXT_COLON.finditer(text):
        _add(m.group(1), m.group(2), 0.90, m.group(0))
    for m in _RE_TXT_EQUALS.finditer(text):
        _add(m.group(1), m.group(2), 0.85, m.group(0))
    for m in _RE_TXT_TABLE.finditer(text):
        _add(m.group(1), m.group(2), 0.75, m.group(0))
 
    return list(seen.values())
 
 
def _txt_llm_extract(text: str) -> list:
    _system = (
        "You are a document field extractor. "
        "Extract every key-value pair from the document text. "
        "Return ONLY valid JSON: "
        '{"fields": [{"field_name": "...", "value": "...", "confidence": 0.0-1.0, "source_text": "..."}]}'
        " No markdown. No preamble. Use labels exactly as written. "
        "confidence: 0.95 if explicitly labelled, 0.75 if inferred."
    )
    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            azure_endpoint=_os_txt.environ.get("OPENAI_DEPLOYMENT_ENDPOINT", ""),
            api_key=_os_txt.environ.get("OPENAI_API_KEY", ""),
            api_version=_os_txt.environ.get("OPENAI_API_VERSION", "2024-12-01-preview"),
        )
        model = _os_txt.environ.get("OPENAI_DEPLOYMENT_NAME", "gpt-4o-mini")
        truncated = text[:6000] + ("\n\n[... truncated ...]" if len(text) > 6000 else "")
        response = client.chat.completions.create(
            model=model, max_tokens=2500, temperature=0.0,
            messages=[
                {"role": "system", "content": _system},
                {"role": "user", "content": f"Extract fields from:\\n\\n{truncated}"},
            ],
        )
        raw = (response.choices[0].message.content or "").strip()
        raw = _re_txt.sub(r"^```(?:json)?\\s*", "", raw)
        raw = _re_txt.sub(r"\\s*```$", "", raw).strip()
        parsed = _json_txt.loads(raw)
        out = []
        for f in parsed.get("fields", []):
            if not f.get("field_name") or not f.get("value"):
                continue
            out.append({
                "field_name": str(f["field_name"]).strip(),
                "value": str(f["value"]).strip(),
                "confidence": float(f.get("confidence", 0.80)),
                "source_text": str(f.get("source_text", ""))[:200],
                "bounding_polygon": None, "source_page": 1,
                "page_width": None, "page_height": None,
                "excel_row": 1, "excel_col": None,
            })
        return out
    except Exception:
        return []
 
 
def _txt_merge(regex_fields: list, llm_fields: list) -> list:
    def _nk(s: str) -> str:
        return _re_txt.sub(r"[\\s_\\-]+", "_", (s or "").strip().lower())
    merged: dict = {}
    for f in llm_fields:
        merged[_nk(f["field_name"])] = f
    for f in regex_fields:
        nk = _nk(f["field_name"])
        if nk not in merged:
            merged[nk] = f
    return list(merged.values())
 
 
def parse_txt_file(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a plain-text / transcript file.
 
    1. Decode bytes to text.
    2. Run generic regex extraction (colon, equals, tabular patterns).
    3. If regex finds < _TXT_MIN_REGEX_FIELDS: call LLM as fallback.
    4. Merge results (LLM priority, regex fills gaps).
 
    Returns dict compatible with run_pdf_intelligence():
      {"full_text", "pages", "doc_type", "doc_label", "source", ...}
    """
    text = ""
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    if not text:
        text = file_bytes.decode("utf-8", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
 
    regex_fields = _txt_regex_extract(text)
    _used_llm = False
 
    if len(regex_fields) < _TXT_MIN_REGEX_FIELDS:
        llm_fields = _txt_llm_extract(text)
        if llm_fields:
            fields = _txt_merge(regex_fields, llm_fields)
            _used_llm = True
        else:
            fields = regex_fields
    else:
        fields = regex_fields
 
    return {
        "full_text":           text,
        "source":              "txt",
        "doc_type":            "txt_document",
        "doc_label":           "Text Document",
        "_used_llm":           _used_llm,
        "_regex_field_count":  len(regex_fields),
        "pages": [{
            "page_num":   1,
            "page_label": "Page 1",
            "raw_text":   text,
            "fields":     fields,
        }],
    }    
